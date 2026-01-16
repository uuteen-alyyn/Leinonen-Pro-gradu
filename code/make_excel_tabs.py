import json
from pathlib import Path
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

QUESTION_ORDER = [
    "A1","A2","A3","A4","A5","A6","A7","A8","A9","A10",
    "B1","B2","B3","B4","B5","B6","B7","B8","B9","B10",
]

PROVIDERS = [
    ("OpenAI", "out_openai.jsonl"),
    ("Claude", "out_claude.jsonl"),
    ("Gemini", "out_gemini.jsonl"),
]

def read_jsonl(path: str) -> Dict[str, Dict[str, Any]]:
    """
    Returns mapping: custom_id -> row dict (the whole JSON object per line)
    """
    data = {}
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Missing file: {path}")
    with p.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            cid = str(obj.get("custom_id", "")).strip()
            if not cid:
                raise ValueError(f"{path}: missing custom_id on line {line_no}")
            data[cid] = obj
    return data

def safe_sheet_name(name: str) -> str:
    # Excel sheet name rules: max 31 chars, cannot contain : \ / ? * [ ]
    bad = [":", "\\", "/", "?", "*", "[", "]"]
    s = name
    for ch in bad:
        s = s.replace(ch, "_")
    return s[:31] if len(s) > 31 else s

def main():
    # Load all providers
    provider_data = {}
    for label, fname in PROVIDERS:
        provider_data[label] = read_jsonl(fname)

    # Union of all article IDs across providers (usually identical)
    all_ids = set()
    for label in provider_data:
        all_ids |= set(provider_data[label].keys())
    all_ids = sorted(all_ids)

    wb = Workbook()
    # Remove default empty sheet
    default = wb.active
    wb.remove(default)

    for article_id in all_ids:
        ws = wb.create_sheet(title=safe_sheet_name(article_id))

        # Header row
        ws["A1"] = "Question"
        for col_i, (provider_label, _fname) in enumerate(PROVIDERS, start=2):
            ws.cell(row=1, column=col_i).value = provider_label

        # Fill question rows
        for r_i, q in enumerate(QUESTION_ORDER, start=2):
            ws.cell(row=r_i, column=1).value = q

            for c_i, (provider_label, _fname) in enumerate(PROVIDERS, start=2):
                row = provider_data.get(provider_label, {}).get(article_id)

                # If article missing for that provider -> leave blank
                if not row:
                    continue

                # If that run errored -> leave blank (you can change to "ERR" if you want)
                if row.get("error"):
                    continue

                ans = (row.get("answers") or {}).get(q, None)
                if ans in (0, 1):
                    ws.cell(row=r_i, column=c_i).value = int(ans)
                elif ans is None:
                    pass
                else:
                    # Sometimes answers might be "0"/"1" as strings
                    try:
                        ws.cell(row=r_i, column=c_i).value = 1 if str(ans).strip() == "1" else 0
                    except Exception:
                        pass

        # Make it readable
        ws.column_dimensions["A"].width = 12
        for c_i in range(2, 2 + len(PROVIDERS)):
            ws.column_dimensions[get_column_letter(c_i)].width = 10
        ws.freeze_panes = "B2"

    out_path = "results_matrix_by_article.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path} with {len(all_ids)} sheets.")

if __name__ == "__main__":
    main()
import json
from pathlib import Path
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

QUESTION_ORDER = [
    "A1","A2","A3","A4","A5","A6","A7","A8","A9","A10",
    "B1","B2","B3","B4","B5","B6","B7","B8","B9","B10",
]

PROVIDERS = [
    ("OpenAI", "out_openai.jsonl"),
    ("Claude", "out_claude.jsonl"),
    ("Gemini", "out_gemini.jsonl"),
]

def read_jsonl(path: str) -> Dict[str, Dict[str, Any]]:
    """
    Returns mapping: custom_id -> row dict (the whole JSON object per line)
    """
    data = {}
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Missing file: {path}")
    with p.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            cid = str(obj.get("custom_id", "")).strip()
            if not cid:
                raise ValueError(f"{path}: missing custom_id on line {line_no}")
            data[cid] = obj
    return data

def safe_sheet_name(name: str) -> str:
    # Excel sheet name rules: max 31 chars, cannot contain : \ / ? * [ ]
    bad = [":", "\\", "/", "?", "*", "[", "]"]
    s = name
    for ch in bad:
        s = s.replace(ch, "_")
    return s[:31] if len(s) > 31 else s

def main():
    # Load all providers
    provider_data = {}
    for label, fname in PROVIDERS:
        provider_data[label] = read_jsonl(fname)

    # Union of all article IDs across providers (usually identical)
    all_ids = set()
    for label in provider_data:
        all_ids |= set(provider_data[label].keys())
    all_ids = sorted(all_ids)

    wb = Workbook()
    # Remove default empty sheet
    default = wb.active
    wb.remove(default)

    for article_id in all_ids:
        ws = wb.create_sheet(title=safe_sheet_name(article_id))

        # Header row
        ws["A1"] = "Question"
        for col_i, (provider_label, _fname) in enumerate(PROVIDERS, start=2):
            ws.cell(row=1, column=col_i).value = provider_label

        # Fill question rows
        for r_i, q in enumerate(QUESTION_ORDER, start=2):
            ws.cell(row=r_i, column=1).value = q

            for c_i, (provider_label, _fname) in enumerate(PROVIDERS, start=2):
                row = provider_data.get(provider_label, {}).get(article_id)

                # If article missing for that provider -> leave blank
                if not row:
                    continue

                # If that run errored -> leave blank (you can change to "ERR" if you want)
                if row.get("error"):
                    continue

                ans = (row.get("answers") or {}).get(q, None)
                if ans in (0, 1):
                    ws.cell(row=r_i, column=c_i).value = int(ans)
                elif ans is None:
                    pass
                else:
                    # Sometimes answers might be "0"/"1" as strings
                    try:
                        ws.cell(row=r_i, column=c_i).value = 1 if str(ans).strip() == "1" else 0
                    except Exception:
                        pass

        # Make it readable
        ws.column_dimensions["A"].width = 12
        for c_i in range(2, 2 + len(PROVIDERS)):
            ws.column_dimensions[get_column_letter(c_i)].width = 10
        ws.freeze_panes = "B2"

    out_path = "results_matrix_by_article.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path} with {len(all_ids)} sheets.")

if __name__ == "__main__":
    main()
